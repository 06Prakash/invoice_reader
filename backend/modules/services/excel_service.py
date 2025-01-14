import pandas as pd
import os
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from modules.logging_util import setup_logger
logger = setup_logger(__name__)
import re

def sanitize_sheet_name(sheet_name):
    return re.sub(r'[\\/*?:\[\]]', '', sheet_name)[:31]

def convert_parentheses_to_negative(df):
    """
    Converts values in parentheses (e.g., "(10)") to negative numbers (e.g., -10).
    Handles values with commas like "(3,140)" as well.
    Restores commas for thousands separators after conversion for readability.
    Ensures integers are stored without decimals.

    Args:
        df (pd.DataFrame): Input DataFrame to process.

    Returns:
        pd.DataFrame: DataFrame with converted values and formatted numbers.
    """
    logger.info("Attempting to convert parentheses to negative numbers..")
    try:
        # Apply conversion with error handling
        def safe_convert(value):
            if isinstance(value, str) and value.startswith('(') and value.endswith(')'):
                try:
                    # Remove commas, strip parentheses, and convert to negative float
                    converted_value = -float(value.replace(',', '').strip('()'))
                    return format_number_with_commas(converted_value)
                except ValueError as e:
                    logger.error(f"Failed to convert value: {value} with error: {e}")
                    return value  # Return original value if conversion fails
            elif isinstance(value, (int, float)):
                # Format numeric values with commas
                return format_number_with_commas(value)
            return value

        def format_number_with_commas(number):
            """
            Formats a number with commas for thousands separators.
            Retains numeric type for compatibility with Excel storage.
            """
            if isinstance(number, (int, float)):
                # Check if the number is effectively an integer
                if isinstance(number, float) and number.is_integer():
                    return "{:,}".format(int(number))  # Format as integer with commas
                elif isinstance(number, float):
                    return "{:,.2f}".format(number)  # Format float with two decimal places
                return "{:,}".format(number)  # Format integer with commas
            return number

        # Apply the conversion safely to the DataFrame
        return df.applymap(safe_convert)
    except Exception as e:
        logger.error(f"Unexpected error in convert_parentheses_to_negative: {e}")
        raise  # Propagate exception for visibility

def remove_columns(df, texts_to_match):
    """
    Removes columns from the DataFrame that match any of the specified texts in the first row.

    Args:
        df (pd.DataFrame): Input DataFrame to process.
        texts_to_match (list): List of texts to match in the first row.

    Returns:
        pd.DataFrame: Processed DataFrame with the matching columns removed.
    """
    try:
        logger.info(f"Attempting to remove columns with any of the texts {texts_to_match} in the first row.")

        # Normalize column names
        df.columns = df.columns.map(str).str.strip().str.lower()

        # Normalize first-row values
        first_row_normalized = df.iloc[0].astype(str).str.strip().str.lower()

        # Normalize texts_to_match
        texts_to_match = [text.strip().lower() for text in texts_to_match if text.strip()]

        # Find matching columns
        matching_columns = first_row_normalized[first_row_normalized.isin(texts_to_match)].index.tolist()
        
        if matching_columns:
            logger.info(f"Found matching columns {matching_columns}. Removing them.")
            df = df.drop(columns=matching_columns, errors='ignore')
        else:
            logger.warning(f"No matching texts {texts_to_match} found in the first row.")

        logger.info(f"DataFrame after column removal:\n{df.head()}")
        return df
    except Exception as e:
        logger.error(f"Error in remove_columns_by_row_values: {e}")
        raise


def consolidate_tables(tables):
    """
    Consolidates a list of DataFrames into a single DataFrame.
    """
    logger.info(f"Attempting to consolidate tables {', '.join(tables)}")
    return pd.concat(tables, ignore_index=True) if tables else pd.DataFrame()


def remove_gridlines(writer, sheet_name):
    """
    Removes gridlines from the specified sheet in the Excel writer.
    """
    logger.info(f"Attempting to remove gridlines from sheet {sheet_name}..")
    worksheet = writer.sheets[sheet_name]
    worksheet.sheet_view.showGridLines = False

def save_sheet(writer, df, sheet_name, config):
    """
    Saves a DataFrame to a specific sheet in the Excel writer with beautification.

    Args:
        writer (ExcelWriter): Excel writer object.
        df (pd.DataFrame): DataFrame to save.
        sheet_name (str): Name of the sheet.
        config (dict): Configuration dictionary.
    """
    safe_sheet_name = sanitize_sheet_name(sheet_name)
    if df.empty:
        logger.warning(f"Skipping empty DataFrame for sheet: {safe_sheet_name}")
        return
    try:
        df.to_excel(writer, sheet_name=safe_sheet_name, index=False, header=False)
        beautify_excel(writer, safe_sheet_name)  # Beautify the sheet
        if config.get("gridLinesRemoval", False) and safe_sheet_name in writer.sheets:
            remove_gridlines(writer, safe_sheet_name)
    except Exception as e:
        logger.error(f"Failed to save sheet '{safe_sheet_name}': {e}")

def remove_rows(df, identity_to_remove_row):
    """
    Removes rows from the DataFrame based on specific identifiers.

    Args:
        df (pd.DataFrame): Input DataFrame to process.
        identity_to_remove_row (list): List of identifiers to remove from the DataFrame.
                                       These identifiers can be present in any column.

    Returns:
        pd.DataFrame: The DataFrame after removing the specified rows.
    """
    if not identity_to_remove_row:
        logger.info("No identifiers provided to remove rows. Returning the original DataFrame.")
        return df

    logger.info(f"Attempting to remove rows with identifiers: {identity_to_remove_row}")
    identity_to_remove_row = [identifier.strip().lower() for identifier in identity_to_remove_row if identifier.strip()]

    if len(identity_to_remove_row) > 0:
        # Normalize the identifiers to a set for faster lookups
        identifiers_set = set(identity_to_remove_row)
        # Apply a filter to keep only rows that do not match any identifier
        filtered_df = df[~df.apply(lambda row: any(item in identifiers_set for item in row.astype(str)), axis=1)]
        logger.info(f"Removed rows. Original shape: {df.shape}, New shape: {filtered_df.shape}")
        return filtered_df
    return df

def process_table_data(df, config):
    """
    Processes table data by applying transformations like:
    - Removing numeric headers
    - Removing specific columns
    - Converting parentheses to negatives
    """

    logger.info(f"Data Frame : {df}")
    df = convert_parentheses_to_negative(df)
    logger.info("Crossed the paranthesis update function")
    logger.info(f"Current config: {config}")

    if columns := config.get("columnsToRemove", []):
        df = remove_columns(df, columns)
    
    if len(config.get("rowsToRemove", [])) > 0:
        rowsToRemove = config.get("rowsToRemove", [])
        df = remove_rows(df, rowsToRemove)
    logger.info(f"Data Frame after processing: {df}")
    return df

def save_sections_to_excel_and_csv(section_data, filename, output_folder, config=None):
    """
    Saves processed section-specific data to both Excel and a single combined CSV file.
    Incorporates section-specific configurations.

    Args:
        section_data (dict): Section data containing tables or other information.
        filename (str): Name of the input file.
        output_folder (str): Path to save the processed Excel and CSV files.
        config (dict, optional): Configuration dictionary containing per-section settings.

    Returns:
        dict: Result dictionary with 'success' or 'failure' status and relevant paths.
    """
    config = config or {}
    base_filename = os.path.splitext(filename)[0]
    excel_path = os.path.join(output_folder, f"{base_filename}_sections_processed.xlsx")
    combined_csv_path = os.path.join(output_folder, f"{base_filename}_combined.csv")
    logger.info(f"====section_data====={section_data}")

    try:
        has_data = False
        combined_dataframes = []

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            for section, content in section_data.items():
                table_content = content.get("raw_tables", None)
                if table_content == None:
                    table_content = content
                section_dfs = process_and_save_section(table_content, section, base_filename, output_folder, writer, config)
                combined_dataframes.extend(section_dfs)
                has_data = True

        if not has_data:
            logger.warning("No valid data found. Adding default placeholder sheet.")
            pd.DataFrame(["No data available"]).to_excel(writer, sheet_name="Placeholder", index=False)

        # Combine all DataFrames into a single CSV
        if combined_dataframes:
            combined_df = pd.concat(combined_dataframes, ignore_index=True)
            combined_df.to_csv(combined_csv_path, index=False)
            logger.info(f"Saved combined CSV at {combined_csv_path}")

        logger.info(f"Processed section data saved to Excel at {excel_path}")
        return {"result": "success", "excel_path": excel_path, "csv_path": combined_csv_path}
    except Exception as e:
        logger.error(f"Failed to save sections to Excel and combined CSV: {e}")
        if excel_path and os.path.exists(excel_path):
            os.remove(excel_path)
        if combined_csv_path and os.path.exists(combined_csv_path):
            os.remove(combined_csv_path)
        return {"result": "failure", "error": str(e)}

def process_and_save_section(content, section, base_filename, output_folder, writer, config):
    """
    Processes a single section and saves its data to both Excel and returns the DataFrame for CSV combination.

    Args:
        content (any): Section content (list, dict, or str).
        section (str): Name of the section.
        base_filename (str): Base name for files.
        output_folder (str): Path to save files.
        writer (ExcelWriter): Excel writer object.
        config (dict): Configuration dictionary.

    Returns:
        list: List of DataFrames to be combined into a single CSV.
    """
    section_dataframes = []
    section_config = config.get(section, {}).get('excel', {})
    columns_to_remove = section_config.get('columnsToRemove', [])
    grid_lines_removal = section_config.get('gridLinesRemoval', False)
    rows_to_remove = section_config.get('rowsToRemove', [])

    logger.info(f"Processing section: {section} with config: {section_config}, Content Type: {type(content)}")

    if isinstance(content, list):  # Process tables
        for idx, table in enumerate(content):
            if isinstance(table, pd.DataFrame) and not table.empty:
                logger.info(f"Writing table {idx + 1} for section {section}, shape: {table.shape}")
                safe_sheet_name = sanitize_sheet_name(f"{section}_Table_{idx + 1}")
                logger.info(f"Sheet name: {safe_sheet_name}")

                # Process table data with section-specific configuration
                table = process_table_data(
                    table,
                    {"columnsToRemove": columns_to_remove, "gridLinesRemoval": grid_lines_removal, 'rowsToRemove': rows_to_remove}
                )
                save_sheet(
                    writer,
                    table,
                    safe_sheet_name,
                    {"gridLinesRemoval": grid_lines_removal}
                )
                table["Section"] = section
                section_dataframes.append(table)
            else:
                logger.warning(f"Skipping empty or invalid table for section: {section}")

    elif isinstance(content, dict):  # Process field-based data
        logger.info(f"Dictionary content: {content}")
        df = pd.DataFrame(content.items(), columns=["Field", "Value"])
        if not df.empty:
            safe_sheet_name = sanitize_sheet_name(section)
            df = process_table_data(
                df,
                {"columnsToRemove": columns_to_remove, "gridLinesRemoval": grid_lines_removal}
            )
            save_sheet(
                writer,
                df,
                safe_sheet_name,
                {"gridLinesRemoval": grid_lines_removal}
            )
            df["Section"] = section
            section_dataframes.append(df)
        else:
            logger.warning(f"Generated empty DataFrame for section: {section}")

    elif isinstance(content, str):  # Process text data
        rows = [row.split() for row in content.split("\n") if row.strip()]
        df = pd.DataFrame(rows)
        if not df.empty:
            safe_sheet_name = sanitize_sheet_name(section)
            df = process_table_data(
                df,
                {"columnsToRemove": columns_to_remove, "gridLinesRemoval": grid_lines_removal}
            )
            save_sheet(
                writer,
                df,
                safe_sheet_name,
                {"gridLinesRemoval": grid_lines_removal}
            )
            df["Section"] = section
            section_dataframes.append(df)
        else:
            logger.warning(f"Generated empty DataFrame for section: {section}")

    else:
        logger.warning(f"Unsupported content type for section: {section}. Skipping.")

    return section_dataframes

def consolidate_excel_sheets(input_files, output_excel_path):
    """
    Consolidates multiple Excel files by matching rows in specified sheets
    and creating a new consolidated Excel file with beautification.

    Args:
        input_files (list): List of input Excel file paths.
        output_excel_path (str): Path to save the consolidated Excel file.

    Returns:
        dict: Status and path to the consolidated file.
    """
    try:
        logger.info(f"Starting consolidation for files: {input_files}")
        sheet_data = {}  # Dictionary to hold data for each sheet across files

        # Read all input files and load the data for each sheet
        for file in input_files:
            workbook = pd.ExcelFile(file)
            for sheet_name in workbook.sheet_names:
                sanitized_name = sanitize_sheet_name(sheet_name)
                if sanitized_name not in sheet_data:
                    sheet_data[sanitized_name] = []
                # Read the sheet and append its DataFrame to the list
                df = workbook.parse(sheet_name)
                logger.info(f"Loaded data for sheet {sanitized_name} from file {file}, shape: {df.shape}")
                sheet_data[sanitized_name].append(df)

        # Consolidate data for each sheet
        with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
            for sheet_name, dataframes in sheet_data.items():
                logger.info(f"Consolidating sheet: {sheet_name}")
                consolidated_df = consolidate_dataframes(dataframes)
                # remove_gridlines(writer, sheet_name)
                if not consolidated_df.empty:
                    consolidated_df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
                    logger.info(f"Consolidated data written to sheet: {sheet_name}")
                    # Apply beautification to the sheet
                    beautify_excel(writer, sheet_name)
                else:
                    logger.warning(f"No data to consolidate for sheet: {sheet_name}")

        logger.info(f"Consolidated Excel file saved at {output_excel_path}")
        return {
            "status": "success",
            "message": f"Consolidated Excel file saved at {output_excel_path}",
            "output_path": output_excel_path,
        }

    except Exception as e:
        logger.error(f"Error during consolidation: {e}")
        return {"status": "error", "message": str(e)}
    
def consolidate_dataframes(dataframes):
    """
    Consolidates a list of DataFrames by matching rows based on the first column.

    Args:
        dataframes (list): List of DataFrames to consolidate.

    Returns:
        pd.DataFrame: Consolidated DataFrame.
    """
    try:
        if not dataframes:
            logger.warning("No DataFrames to consolidate.")
            return pd.DataFrame()

        # Use the first DataFrame as the base
        consolidated_df = dataframes[0].copy()
        key_column = consolidated_df.columns[0]  # Assume the first column is the key

        for df in dataframes[1:]:
            # Merge DataFrames on the key column
            df = df.copy()
            consolidated_df = pd.merge(
                consolidated_df,
                df,
                on=key_column,
                how="outer",
                suffixes=("", "_other")
            )

            # Sum numeric columns for matching rows
            numeric_cols = consolidated_df.select_dtypes(include="number").columns
            for col in numeric_cols:
                if col.endswith("_other"):
                    base_col = col.replace("_other", "")
                    if base_col in consolidated_df:
                        consolidated_df[base_col] += consolidated_df[col]
                        consolidated_df.drop(columns=[col], inplace=True)

            # Fill NaN with 0 for numeric columns
            consolidated_df[numeric_cols] = consolidated_df[numeric_cols].fillna(0)

        return consolidated_df

    except Exception as e:
        logger.error(f"Error during DataFrame consolidation: {e}")
        return pd.DataFrame()

def beautify_excel(writer, sheet_name, max_column_width=50):
    """
    Beautifies the specified Excel sheet.
    - Formats headers with bold fonts, centered alignment, and a background color.
    - Applies word wrapping to all cells.
    - Restricts column widths to a maximum value.

    Args:
        writer (ExcelWriter): Excel writer object.
        sheet_name (str): Name of the sheet to beautify.
        max_column_width (int, optional): Maximum allowed column width. Default is 50.
    """
    try:
        logger.info(f"Applying beautification to sheet: {sheet_name}")
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Apply header formatting
        for row in worksheet.iter_rows(min_row=1, max_row=1):  # Format only the header row
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment

        # Format content rows and adjust column widths
        for col_idx, column_cells in enumerate(worksheet.columns, start=1):
            # Calculate max content length
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells
            )
            adjusted_width = min(max_length + 2, max_column_width)  # Cap column width at max_column_width
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = adjusted_width

            # Apply word wrapping for all cells in the column
            for cell in column_cells:
                if cell.value:  # Apply alignment only if the cell has a value
                    cell.alignment = Alignment(wrap_text=True)

        logger.info(f"Beautification applied to sheet: {sheet_name} with max column width {max_column_width}")
    except Exception as e:
        logger.error(f"Error beautifying sheet '{sheet_name}': {e}")


def consolidate_and_save(writer, consolidated_data, config):
    """
    Consolidates multiple DataFrames and saves them to a single sheet with beautification.

    Args:
        writer (ExcelWriter): Excel writer object.
        consolidated_data (list): List of DataFrames to consolidate.
        config (dict): Configuration dictionary.
    """
    if consolidated_data:
        merged_df = consolidate_tables(consolidated_data)
        sheet_name = "Consolidated"
        save_sheet(writer, merged_df, sheet_name, config)
        beautify_excel(writer, sheet_name)  # Beautify the consolidated sheet
