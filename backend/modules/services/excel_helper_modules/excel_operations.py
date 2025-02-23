import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from modules.services.excel_helper_modules.sanitization import sanitize_sheet_name
from modules.logging_util import setup_logger

logger = setup_logger(__name__)

def save_sheet(writer, df, sheet_name, config):
    """
    Saves a DataFrame to a specific sheet in the Excel writer with beautification.

    Args:
        writer (ExcelWriter): Excel writer object.
        df (pd.DataFrame): DataFrame to save.
        sheet_name (str): Name of the sheet.
        config (dict): Configuration dictionary.
    """
    safe_sheet_name = sanitize_sheet_name(sheet_name)
    if df.empty:
        logger.warning(f"Skipping empty DataFrame for sheet: {safe_sheet_name}")
        return
    try:
        # Convert headers to uppercase
        df.columns = [col.upper() for col in df.columns]
        df.to_excel(writer, sheet_name=safe_sheet_name, index=False, header=True)
        beautify_excel(writer, safe_sheet_name)  # Beautify the sheet
        if config.get("gridLinesRemoval", False) and safe_sheet_name in writer.sheets:
            remove_gridlines(writer, safe_sheet_name)
    except Exception as e:
        logger.error(f"Failed to save sheet '{safe_sheet_name}': {e}")

def beautify_excel(writer, sheet_name, max_column_width=50):
    """
    Beautifies the specified Excel sheet.
    - Formats headers with bold fonts, centered alignment, and a background color.
    - Applies word wrapping to all cells.
    - Restricts column widths to a maximum value.

    Args:
        writer (ExcelWriter): Excel writer object.
        sheet_name (str): Name of the sheet to beautify.
        max_column_width (int, optional): Maximum allowed column width. Default is 50.
    """
    try:
        logger.info(f"Applying beautification to sheet: {sheet_name}")
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Apply header formatting
        for row in worksheet.iter_rows(min_row=1, max_row=1):  # Format only the header row
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment

        # Format content rows and adjust column widths
        for col_idx, column_cells in enumerate(worksheet.columns, start=1):
            # Calculate max content length
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells
            )
            adjusted_width = min(max_length + 2, max_column_width)  # Cap column width at max_column_width
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = adjusted_width

            # Apply word wrapping for all cells in the column
            for cell in column_cells:
                if cell.value:  # Apply alignment only if the cell has a value
                    cell.alignment = Alignment(wrap_text=True)

        logger.info(f"Beautification applied to sheet: {sheet_name} with max column width {max_column_width}")
    except Exception as e:
        logger.error(f"Error beautifying sheet '{sheet_name}': {e}")

def remove_gridlines(writer, sheet_name):
    """
    Removes gridlines from the specified sheet in the Excel writer.
    """
    logger.info(f"Attempting to remove gridlines from sheet {sheet_name}..")
    worksheet = writer.sheets[sheet_name]
    worksheet.sheet_view.showGridLines = False